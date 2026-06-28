from io import BytesIO

from openpyxl import Workbook

from backend.supplier_products.workbook import parse_zvezda_workbook


def test_parse_zvezda_workbook_products_mappings_and_stocks() -> None:
    workbook = Workbook()
    price = workbook.active
    price.title = "Прайс Звезда"
    price.append(
        [
            "Артикул производителя",
            "Наименование",
            "Описание",
            "Картинки",
            "Штрих-код",
            "Кол-во в 1 кор",
            "Вес, гр",
            "Размер, мм",
            "Цена закупки",
            "ЗАКАЗ",
            "Сумма",
        ]
    )
    price.append(
        [
            "3645ГП",
            "3645 Американский танк",
            "L=21,6 см",
            "фото",
            "4600327836458",
            8,
            571,
            "400х242х70",
            "1 494,35",
            0,
            0,
        ]
    )
    price["D2"].hyperlink = "https://zvezda.org.ru/catalog/test-product/"

    catalog = workbook.create_sheet("Справочник Звезда")
    catalog.append(
        [
            "Наименование",
            "Артикул производителя",
            "Артикул WB",
            "Штук в наборе",
            "Артикул продавца",
            "Цена закупки",
            "МРЦ",
        ]
    )
    catalog.append(["Танк", "3645ГП", "123456789", 1, "3645-WB", 1494.35, 2190])

    cards = workbook.create_sheet("Справочник_карточек_WB")
    cards.append(
        [
            "Бренд",
            "Предмет",
            "Код размера (chrt_id)",
            "Артикул продавца",
            "Артикул WB",
            "Размер",
            "Баркод",
            "Объем, л.",
            "Состав",
        ]
    )
    cards.append(["Звезда", "Модель", "1", "3645-WB", "123456789", "", "4600327836458", "", ""])

    stocks = workbook.create_sheet("WB Остатки")
    stocks.append(
        [
            "Дата снимка",
            "Артикул WB",
            "Бренд",
            "Предмет",
            "Артикул продавца",
            "Остаток на складах",
            "В пути до клиента",
            "В пути возвраты",
        ]
    )
    stocks.append(["2026-06-28", "123456789", "Звезда", "Модель", "3645-WB", 12, 2, 1])

    content = BytesIO()
    workbook.save(content)

    products, mappings, snapshots = parse_zvezda_workbook(content.getvalue(), "zvezda")

    assert len(products) == 1
    assert products[0].sku == "3645ГП"
    assert products[0].source_url is not None
    assert products[0].pack_units == 8
    assert products[0].weight_grams == 571
    assert products[0].dimensions == "400х242х70"
    assert len(mappings) == 2
    assert any(mapping.manufacturer_article == "3645ГП" and mapping.wb_article == "123456789" for mapping in mappings)
    assert any(mapping.barcode == "4600327836458" for mapping in mappings)
    assert len(snapshots) == 1
    assert snapshots[0].wb_article == "123456789"
    assert snapshots[0].stock_qty == 12
