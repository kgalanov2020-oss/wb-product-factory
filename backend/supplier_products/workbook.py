from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from .models import SupplierProductInput, WBCardMappingInput, WBStockSnapshotInput
from .parser import parse_xlsx_price_list


def parse_zvezda_workbook(content: bytes, supplier: str) -> tuple[
    list[SupplierProductInput],
    list[WBCardMappingInput],
    list[WBStockSnapshotInput],
]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    products = parse_xlsx_price_list(content, supplier)
    mappings = _parse_mappings(workbook, supplier)
    stocks = _parse_stocks(workbook)
    return products, mappings, stocks


def _parse_mappings(workbook: Any, supplier: str) -> list[WBCardMappingInput]:
    mappings: list[WBCardMappingInput] = []
    if "Справочник Звезда" in workbook.sheetnames:
        ws = workbook["Справочник Звезда"]
        headers = [str(value or "").strip() for value in next(ws.iter_rows(values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = _row(headers, row)
            manufacturer_article = _text(values.get("Артикул производителя"))
            wb_article = _text(values.get("Артикул WB"))
            seller_article = _text(values.get("Артикул продавца"))
            if not manufacturer_article and not wb_article and not seller_article:
                continue
            mappings.append(
                WBCardMappingInput(
                    supplier=supplier,
                    manufacturer_article=manufacturer_article,
                    seller_article=seller_article,
                    wb_article=wb_article,
                    name=_text(values.get("Наименование")),
                    purchase_price=_decimal(values.get("Цена закупки")),
                    retail_price=_decimal(values.get("МРЦ")),
                    pack_units=_int(values.get("Штук в наборе")),
                    raw=values,
                )
            )
    if "Справочник_карточек_WB" in workbook.sheetnames:
        ws = workbook["Справочник_карточек_WB"]
        headers = [str(value or "").strip() for value in next(ws.iter_rows(values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = _row(headers, row)
            wb_article = _text(values.get("Артикул WB"))
            seller_article = _text(values.get("Артикул продавца"))
            barcode = _text(values.get("Баркод"))
            if not wb_article and not seller_article and not barcode:
                continue
            mappings.append(
                WBCardMappingInput(
                    supplier=supplier,
                    seller_article=seller_article,
                    wb_article=wb_article,
                    barcode=barcode,
                    brand=_text(values.get("Бренд")),
                    subject=_text(values.get("Предмет")),
                    raw=values,
                )
            )
    return mappings


def _parse_stocks(workbook: Any) -> list[WBStockSnapshotInput]:
    if "WB Остатки" not in workbook.sheetnames:
        return []
    ws = workbook["WB Остатки"]
    headers = [str(value or "").strip() for value in next(ws.iter_rows(values_only=True))]
    stocks: list[WBStockSnapshotInput] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = _row(headers, row)
        wb_article = _text(values.get("Артикул WB"))
        if not wb_article:
            continue
        stocks.append(
            WBStockSnapshotInput(
                wb_article=wb_article,
                seller_article=_text(values.get("Артикул продавца")),
                brand=_text(values.get("Бренд")),
                subject=_text(values.get("Предмет")),
                stock_qty=_int(values.get("Остаток на складах")) or 0,
                in_way_to_client=_int(values.get("В пути до клиента")) or 0,
                in_way_from_client=_int(values.get("В пути возвраты")) or 0,
                raw=values,
            )
        )
    return stocks


def _row(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        headers[index]: _json_value(value)
        for index, value in enumerate(row)
        if index < len(headers) and headers[index]
    }


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _int(value: Any) -> int | None:
    decimal = _decimal(value)
    return int(decimal) if decimal is not None else None
