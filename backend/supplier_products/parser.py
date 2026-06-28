from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from .exceptions import SupplierPriceListError
from .models import SupplierProductInput


PHOTO_PATTERN = re.compile(r"https?://[^\s,;\"']+\.(?:jpg|jpeg|png|webp)", re.IGNORECASE)
URL_PATTERN = re.compile(r"https?://[^\s,;\"']+", re.IGNORECASE)

COLUMN_ALIASES = {
    "sku": ("артикул", "код", "sku", "id", "номенклатура"),
    "barcode": ("штрихкод", "barcode", "ean"),
    "name": ("наименование", "название", "товар", "name", "product"),
    "category": ("категория", "раздел", "category", "group", "группа"),
    "wholesale_price": ("опт", "оптовая", "закуп", "закупочная", "цена", "price"),
    "retail_price": ("розница", "ррц", "retail"),
    "stock": ("остаток", "наличие", "stock", "qty", "кол-во", "количество"),
    "pack_units": ("кол-во в 1 кор", "штук в короб", "кратность"),
    "weight_grams": ("вес", "гр"),
    "dimensions": ("размер", "габарит"),
    "description": ("описание",),
    "order_quantity": ("заказ",),
    "source_url": ("ссылка", "url", "страница"),
    "photo_urls": ("фото", "картинка", "картинки", "image", "photo", "изображение"),
}


def parse_csv_price_list(content: bytes, supplier: str) -> list[SupplierProductInput]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = csv.DictReader(io.StringIO(text), dialect=dialect)
    return _parse_dict_rows(list(rows), supplier)


def parse_xlsx_price_list(content: bytes, supplier: str) -> list[SupplierProductInput]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    sheet = _select_sheet(workbook)
    rows = list(sheet.iter_rows())
    if not rows:
        return []
    headers = _headers_for_sheet(sheet.title, rows[0])
    dict_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        values: dict[str, Any] = {}
        for index, cell in enumerate(row):
            if index >= len(headers):
                continue
            header = headers[index]
            value = cell.value
            if cell.hyperlink and cell.hyperlink.target:
                value = cell.hyperlink.target
            values[header] = value
        dict_rows.append(values)
    return _parse_dict_rows(dict_rows, supplier)


def _select_sheet(workbook: Any) -> Any:
    preferred = ("Прайс Звезда", "Справочник Звезда", "Новинки для запуска")
    for title in preferred:
        if title in workbook.sheetnames:
            return workbook[title]
    return workbook.active


def _headers_for_sheet(title: str, row: Any) -> list[str]:
    if title == "Прайс Звезда":
        return [
            "sku",
            "name",
            "description",
            "source_url",
            "barcode",
            "pack_units",
            "weight_grams",
            "dimensions",
            "wholesale_price",
            "order_quantity",
            "total",
        ]
    headers: list[str] = []
    for index, cell in enumerate(row):
        value = str(cell.value or "").strip()
        headers.append(value or f"column_{index + 1}")
    return headers


def _parse_dict_rows(rows: list[dict[str, Any]], supplier: str) -> list[SupplierProductInput]:
    products: list[SupplierProductInput] = []
    for row in rows:
        normalized = {_normalize_key(key): value for key, value in row.items()}
        name = _first_value(normalized, "name")
        if not name:
            continue
        photo_urls = _extract_urls(_first_value(normalized, "photo_urls"), only_images=True)
        source_urls = _extract_urls(_first_value(normalized, "source_url"), only_images=False)
        for value in row.values():
            photo_urls.extend(url for url in _extract_urls(value, only_images=True) if url not in photo_urls)
        try:
            products.append(
                SupplierProductInput(
                    supplier=supplier,
                    sku=_as_text(_first_value(normalized, "sku")),
                    barcode=_as_text(_first_value(normalized, "barcode")),
                    name=str(name).strip(),
                    category=_as_text(_first_value(normalized, "category")),
                    wholesale_price=_as_decimal(_first_value(normalized, "wholesale_price")),
                    retail_price=_as_decimal(_first_value(normalized, "retail_price")),
                    stock=_as_int(_first_value(normalized, "stock")),
                    pack_units=_as_int(_first_value(normalized, "pack_units")),
                    weight_grams=_as_decimal(_first_value(normalized, "weight_grams")),
                    dimensions=_as_text(_first_value(normalized, "dimensions")),
                    description=_as_text(_first_value(normalized, "description")),
                    order_quantity=_as_int(_first_value(normalized, "order_quantity")),
                    photo_urls=photo_urls,
                    source_url=source_urls[0] if source_urls else None,
                    raw={str(key): value for key, value in row.items()},
                )
            )
        except ValueError:
            continue
    return products


def _normalize_key(key: str) -> str:
    lowered = str(key or "").strip().lower()
    for canonical, aliases in COLUMN_ALIASES.items():
        if any(alias in lowered for alias in aliases):
            return canonical
    return lowered


def _first_value(row: dict[str, Any], key: str) -> Any:
    value = row.get(key)
    if value is not None and str(value).strip():
        return value
    return None


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _as_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def _as_int(value: Any) -> int | None:
    decimal = _as_decimal(value)
    return int(decimal) if decimal is not None else None


def _extract_urls(value: Any, only_images: bool) -> list[str]:
    if value is None:
        return []
    text = str(value)
    pattern = PHOTO_PATTERN if only_images else URL_PATTERN
    return [match.group(0) for match in pattern.finditer(text)]


def parse_price_list(content: bytes, filename: str, supplier: str) -> list[SupplierProductInput]:
    lowered = filename.lower()
    if lowered.endswith(".xlsx"):
        return parse_xlsx_price_list(content, supplier)
    if lowered.endswith(".csv") or lowered.endswith(".txt"):
        return parse_csv_price_list(content, supplier)
    raise SupplierPriceListError("Unsupported price list format. Use CSV or XLSX.")
